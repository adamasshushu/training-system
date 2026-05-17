"""报告导出路由：Excel格式导出学习报告"""
import io
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from app.database import get_db
from app.models.user import User
from app.models.department import Department
from app.models.course import Course, CourseProgress, Chapter, Lesson
from app.models.exam import Exam, ExamResult
from app.utils.auth import get_current_user

router = APIRouter(prefix="/api/reports", tags=["报告导出"])


@router.get("/export")
async def export_report(
    report_type: str = Query("overview", alias="type", description="overview/department/course"),
    department_id: int = Query(None, alias="department_id"),
    course_id: int = Query(None, alias="course_id"),
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """导出学习报告为Excel"""
    if current_user.get("角色") != "admin":
        raise HTTPException(status_code=403, detail="仅管理员可导出报告")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="导出功能不可用(缺少openpyxl)")

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    def style_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def auto_width(ws, data):
        for col in range(1, len(data[0]) + 1):
            max_len = max(len(str(row[col - 1] or "")) for row in data)
            ws.column_dimensions[chr(64 + col)].width = min(max(max_len * 2, 10), 40)

    # === 概览报表 ===
    ws1 = wb.active
    ws1.title = "学习概览"
    headers1 = ["部门", "员工数", "总课程数", "总学习记录", "总考试记录"]
    style_header(ws1, headers1)

    # 数据
    dept_result = await db.execute(
        select(Department).where(Department.is_active == True).order_by(Department.sort)
    )
    departments = dept_result.scalars().all()

    rows1 = []
    total_users = 0
    total_progress = 0
    total_exams = 0

    for dept in departments:
        users_in_dept = await db.execute(
            select(User).where(User.department_id == dept.id, User.is_active == True)
        )
        dept_users = users_in_dept.scalars().all()
        user_count = len(dept_users)
        total_users += user_count

        prog_result = await db.execute(
            select(CourseProgress).join(User).where(User.department_id == dept.id)
        )
        prog_count = len(prog_result.scalars().all())
        total_progress += prog_count

        exam_r = await db.execute(
            select(ExamResult).join(User).where(User.department_id == dept.id)
        )
        exam_count = len(exam_r.scalars().all())
        total_exams += exam_count

        rows1.append([dept.name, user_count, "", prog_count, exam_count])

    # 合计行
    rows1.append(["合计", total_users, "", total_progress, total_exams])
    for row_data in rows1:
        ws1.append(row_data)
    auto_width(ws1, [headers1] + rows1)

    # === 学员进度报表 ===
    ws2 = wb.create_sheet("学员进度")
    headers2 = ["姓名", "部门", "角色", "邮箱", "已完成课时", "进行中课时", "已完成考试"]
    style_header(ws2, headers2)

    all_users_result = await db.execute(
        select(User).where(User.is_active == True).order_by(User.id)
    )
    all_users = all_users_result.scalars().all()

    rows2 = []
    for u in all_users:
        dept_name = ""
        if u.department_id:
            d = (await db.execute(select(Department).where(Department.id == u.department_id))).scalar_one_or_none()
            if d: dept_name = d.name

        completed = await db.execute(
            select(CourseProgress).where(
                CourseProgress.user_id == u.id,
                CourseProgress.completed == True,
            )
        )
        completed_count = len(completed.scalars().all())

        in_progress = await db.execute(
            select(CourseProgress).where(
                CourseProgress.user_id == u.id,
                CourseProgress.completed == False,
                CourseProgress.progress > 0,
            )
        )
        in_progress_count = len(in_progress.scalars().all())

        exam_done = await db.execute(
            select(ExamResult).where(ExamResult.user_id == u.id)
        )
        exam_done_count = len(exam_done.scalars().all())

        role_map = {"admin": "管理员", "teacher": "讲师", "student": "学员"}
        rows2.append([u.real_name, dept_name, role_map.get(u.role, u.role), u.email or "", completed_count, in_progress_count, exam_done_count])

    for row_data in rows2:
        ws2.append(row_data)
    auto_width(ws2, [headers2] + rows2)

    # === 课程进度报表 ===
    ws3 = wb.create_sheet("课程进度")
    headers3 = ["课程", "分类", "讲师", "总课时", "总学习人数", "已完成人数"]
    style_header(ws3, headers3)

    course_result = await db.execute(
        select(Course).where(Course.is_published == True).order_by(Course.id)
    )
    courses = course_result.scalars().all()

    rows3 = []
    for c in courses:
        cat_name = ""
        if c.category_id:
            cat = (await db.execute(select(CourseCategory).where(CourseCategory.id == c.category_id))).scalar_one_or_none()
            if cat: cat_name = cat.name
        teacher_name = ""
        if c.teacher_id:
            t = (await db.execute(select(User).where(User.id == c.teacher_id))).scalar_one_or_none()
            if t: teacher_name = t.real_name

        total_lessons = await db.execute(
            select(Lesson).join(Chapter).where(Chapter.course_id == c.id)
        )
        lesson_count = len(total_lessons.scalars().all())

        learners = await db.execute(
            select(CourseProgress).where(CourseProgress.course_id == c.id).distinct(CourseProgress.user_id)
        )
        learner_count = len(learners.scalars().all())

        completers = await db.execute(
            select(CourseProgress).where(
                CourseProgress.course_id == c.id,
                CourseProgress.completed == True,
            ).distinct(CourseProgress.user_id)
        )
        completer_count = len(completers.scalars().all())

        from app.models.course import CourseCategory
        rows3.append([c.title, cat_name, teacher_name, lesson_count, learner_count, completer_count])

    for row_data in rows3:
        ws3.append(row_data)
    auto_width(ws3, [headers3] + rows3)

    # 输出Excel
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=training_report.xlsx"},
    )
